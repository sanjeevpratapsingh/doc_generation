import os.path
from docx import Document # Importing Python-Docx which will help in generating MS Doc Files
from docxtpl import DocxTemplate 
# For Playing Templates and creating dynamic templates If needed. Here we can JINJA2 for templating

import openpyxl # importing openpyxl module - For Accessing EXCEL Files and doing operations
import re #For Removing Unwanted Special Characters
  

# Creating Document Object
# document  = Document()

#Create a template with atleast a words other it will give you an error "Package not found at docxtpl"
document = DocxTemplate("template.docx")

#Basic

filepath = 'C:/projects/python/docs/docfiles/'

# Give the location of the file 
#path = "C:\\Users\\Admin\\Desktop\\demo.xlsx" or if file is in the same folder then no need pass the file name
# Creating workbook object and passing the file inside it

def passing_context(the_titles):
    context = { 'title' : the_titles }
    document.render(context)
    document.save(filepath + the_titles+'.docx')

wb_obj = openpyxl.load_workbook("coursetitles.xlsx")
  
sheet_obj = wb_obj.active 
max_row = sheet_obj.max_row 
  
# Loop will print all columns name 
for i in range(2, max_row + 1): 
    row_obj = sheet_obj.cell(row = i, column = 3) 
    row_values = row_obj.value
    the_titles = re.sub(r'[^a-zA-Z0-9 \n\.]', '', str(row_values))
    passing_context(the_titles)
    print(row_values) 


