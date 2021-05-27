import os.path
from docx import Document
from docx.shared import Pt,Cm
from docx.enum.style import WD_STYLE_TYPE

# importing openpyxl module 
import openpyxl 
import re
  
# Creating Document Object
document  = Document()
styles = document.styles
font = document.styles['Normal'].font
font.name = "Poppins"
font.size = Pt(14)


sections = document.sections
for section in sections:
    margin_size = 2
    section.left_margin = Cm(margin_size)
    section.right_margin = Cm(margin_size)
    section.top_margin = Cm(margin_size)
    section.bottom_margin = Cm(margin_size)

filepath = 'C:/projects/python/docs/docfiles/'

# Give the location of the file 
#path = "C:\\Users\\Admin\\Desktop\\demo.xlsx" or if file is in the same folder then no need pass the file name
# Creating workbook object and passing the file inside it

wb_obj = openpyxl.load_workbook("test.xlsx") 
  
sheet_obj = wb_obj.active 
max_row = sheet_obj.max_row 
  
# Loop will print all columns name 
for i in range(2, max_row + 1): 
    row_obj = sheet_obj.cell(row = i, column = 4) 
    row_values = row_obj.value
    the_titles = re.sub(r'[^a-zA-Z0-9 \n\.]', '', row_values)
    document.save(filepath + the_titles+'.docx')
    print(row_values) 


